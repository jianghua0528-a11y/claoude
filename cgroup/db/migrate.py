"""
迁移引导  ·  migrate.py
一次性把旧主文件(xlsx)的字典 + 历史报单 导进新库(Postgres/SQLite)。
部署时跑一次: python -m cgroup.db.migrate /path/to/主文件.xlsx
之后机器人直接往库里写新单, 不再依赖 xlsx。
"""
import re
import sys
from openpyxl import load_workbook
from .session import init_db, get_session
from .models import Broker, Artist, Mama, MamaAssistant, Venue

BROKERS = [("阿星", 0.07, True), ("阿豪", 0.07, True),
           ("老方", 0.07, True), ("老杨", 0.10, False)]

_SPLIT = re.compile(r"[,，、/]+")
_PAREN = re.compile(r"[（(].*?[）)]")


def clean_names(s):
    if not s:
        return []
    out = []
    for part in _SPLIT.split(str(s)):
        name = _PAREN.sub("", part).strip()      # 去掉 (助理)/(惠子) 等括注
        if name:
            out.append(name)
    return out


def migrate(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    init_db()
    s = get_session()

    # 经纪人
    bmap = {}
    for name, pct, share in BROKERS:
        b = s.query(Broker).filter_by(name=name).one_or_none()
        if not b:
            b = Broker(name=name, pct=pct, is_shareholder=share)
            s.add(b); s.flush()
        bmap[name] = b.id

    # 艺人字典
    aw = wb["📋 艺人字典"]
    n_art = 0
    for r in range(4, aw.max_row + 1):
        name = aw.cell(row=r, column=2).value
        if not name:
            continue
        name = str(name).strip()
        if s.query(Artist).filter_by(name=name).first():
            continue
        alias = aw.cell(row=r, column=3).value
        broker = aw.cell(row=r, column=4).value
        s.add(Artist(name=name,
                     aliases=str(alias).strip() if alias and alias != "-" else None,
                     broker_id=bmap.get(str(broker).strip()) if broker else None))
        n_art += 1

    # 妈咪团队架构 (+ 助理)
    mw = wb["📋 妈咪团队架构"]
    n_mama = n_asst = 0
    for r in range(4, mw.max_row + 1):
        name = mw.cell(row=r, column=2).value
        if not name:
            continue
        name = str(name).strip()
        if s.query(Mama).filter_by(name=name).first():
            continue
        status = (mw.cell(row=r, column=3).value or "标准").strip()
        if status == "未合作":
            HIST_DIRECT_MAMAS.add(name)      # 历史直结妈咪(公司0%), 迁移时据此定档
        alias = mw.cell(row=r, column=4).value
        team = mw.cell(row=r, column=5).value
        cycle = (mw.cell(row=r, column=7).value or "半月结").strip()
        m = Mama(name=name, settlement_cycle=cycle,
                 aliases=str(alias).strip() if alias else None)
        s.add(m); s.flush(); n_mama += 1
        for a in clean_names(team):
            s.add(MamaAssistant(name=a, mama_id=m.id)); n_asst += 1

    # 场所字典
    vw = wb["📋 场所字典"]
    n_ven = 0
    for r in range(4, vw.max_row + 1):
        name = vw.cell(row=r, column=2).value
        if not name:
            continue
        name = str(name).strip()
        if s.query(Venue).filter_by(name=name).first():
            continue
        alias = vw.cell(row=r, column=3).value
        vtype = vw.cell(row=r, column=4).value
        ticket = vw.cell(row=r, column=5).value
        rooms = vw.cell(row=r, column=6).value
        try:
            ticket = float(ticket) if ticket not in (None, "") else 0.0
        except (ValueError, TypeError):
            ticket = 0.0          # 备注文字混进门票列 → 当0, 录单时人工补
        s.add(Venue(name=name,
                    aliases=str(alias).strip() if alias and alias != "-" else None,
                    vtype=str(vtype).strip() if vtype else "标准",
                    default_ticket=ticket,
                    rooms=str(rooms).strip() if rooms else None))
        n_ven += 1

    s.commit()
    print(f"字典迁移完成: 经纪人{len(BROKERS)} 艺人{n_art} 妈咪{n_mama}(助理{n_asst}) 场所{n_ven}")

    n_ord = migrate_orders(wb, s)
    s.commit()
    print(f"报单迁移完成: {n_ord} 单 (废单已剔除)")
    s.close()


# ─────────────── 报单迁移 (已验证: 5月挂账/门票精确对上旧库) ───────────────
VOID_KW = ["作废", "退单", "无效", "取消"]
ALIAS_FIX = {"周研": "周妍"}      # 已知笔误
HIST_DIRECT_MAMAS = set()         # 历史"未合作"妈咪名(迁移时填), 其单定档=直结(公司0%)


def _parse_date(dv):
    from datetime import date
    s = str(dv)
    mm = re.search(r"(\d{1,2})月(\d{1,2})日?", s)
    if mm:
        return date(2026, int(mm.group(1)), int(mm.group(2)))
    try:
        return dv.date()
    except Exception:
        return None


def _build_mama_lookup(session):
    from .models import Mama, MamaAssistant
    id_by = {m.name: m.id for m in session.query(Mama).all()}
    alias = {}
    for m in session.query(Mama).all():
        for a in clean_names(m.aliases):
            alias.setdefault(a, m.name)
    for asst in session.query(MamaAssistant).all():
        nm = session.get(Mama, asst.mama_id).name
        alias.setdefault(asst.name, nm)
    for k, v in ALIAS_FIX.items():
        alias[k] = v
    return id_by, alias


def migrate_orders(wb, session):
    from .models import Order, Artist, Venue
    from ..core.status import derive_status
    db = wb["📥 数据库"]
    hdr = {db.cell(row=3, column=c).value: c
           for c in range(1, db.max_column + 1) if db.cell(row=3, column=c).value}
    def g(r, name):
        c = hdr.get(name); return db.cell(row=r, column=c).value if c else None

    mama_id, alias = _build_mama_lookup(session)
    art_id = {a.name: a.id for a in session.query(Artist).all()}
    ven = session.query(Venue).all()
    ven_id = {v.name: v.id for v in ven}
    ven_alias = {}
    for v in ven:
        for a in clean_names(v.aliases):
            ven_alias.setdefault(a, v.id)

    def mode_of(mama, dtype):
        if dtype == "不抽":
            return "全归艺人"
        if not mama or not str(mama).strip():
            return "自单"
        n = str(mama).strip()
        main = n if n in mama_id else alias.get(n, n)
        if main in HIST_DIRECT_MAMAS:       # 历史直结妈咪(公司0%)
            return "直结"
        return "标准"

    n = 0
    for r in range(4, db.max_row + 1):
        if g(r, "序号") is None:
            continue
        dtype = str(g(r, "单据类型") or ""); note = str(g(r, "备注") or "")
        if dtype == "废单" or any(k in note for k in VOID_KW):
            continue
        K = float(g(r, "挂账金额") or 0)
        cm = str(g(r, "现金币种") or "MYR"); Mraw = g(r, "现金金额") or 0
        wpv = g(r, "工价(MYR)"); wpv = float(wpv) if wpv else None
        if cm in ("USDT", "RMB") and Mraw:          # 外币现金: MYR值=工价
            M = wpv or 0.0
            fx_rate = (float(Mraw) / wpv) if wpv else None
        else:
            M = float(Mraw) if cm == "MYR" else 0.0
            fx_rate = None
        mama = g(r, "妈咪")
        mname = str(mama).strip() if mama else None
        mid = mama_id.get(mname) or mama_id.get(alias.get(mname, ""))
        vname = str(g(r, "场所")).strip() if g(r, "场所") else None
        vid = ven_id.get(vname) or ven_alias.get(vname)
        flow = g(r, "[废]M_流向"); flow = str(flow).strip() if flow and str(flow) != "清" else None
        mode = mode_of(mama, dtype)
        cs, ms = derive_status(credit_k=K, cash_m=M, mode=mode, flow=flow)
        session.add(Order(
            seq=g(r, "序号"), biz_date=_parse_date(g(r, "日期")),
            artist_id=art_id.get(str(g(r, "艺名")).strip()) if g(r, "艺名") else None,
            venue_id=vid, room=str(g(r, "包厢")).strip() if g(r, "包厢") else None,
            booker=str(g(r, "预约人")).strip() if g(r, "预约人") else None,
            mama_id=mid, mode=mode, flow=flow,
            credit_k=K, cash_m=M, ticket_o=float(g(r, "门票(MYR)") or 0),
            credit_status=cs, cash_status=ms,
            wp=wpv, currency_k="MYR", currency_m=cm,
            settle_rate=fx_rate,
            customer=str(g(r, "客人")).strip()[:20] if g(r, "客人") else None,
            start_time=str(g(r, "上班时间") or "")[:8], end_time=str(g(r, "下班时间") or "")[:8],
            remark=(note + (f" |原币{Mraw}{cm}" if fx_rate else ""))[:500] or None,
            status="已审核", source_msg_id=f"hist#{g(r,'序号')}"))
        n += 1
    return n


if __name__ == "__main__":
    migrate(sys.argv[1] if len(sys.argv) > 1 else "master.xlsx")
